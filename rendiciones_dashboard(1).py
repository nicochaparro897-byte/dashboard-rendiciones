import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from openpyxl import load_workbook
import os
import base64
import tempfile
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="GCS-P | Dashboard Rendiciones",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── USUARIOS ─────────────────────────────────────────────────────────────────
USUARIOS = {
    "admin":   {"password": "gcsP2026", "rol": "admin",  "nombre": "Administrador"},
    "gerente": {"password": "gcsP2026", "rol": "viewer", "nombre": "Gerencia"},
}

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.main { background-color: #F7F8FC; }
[data-testid="stAppViewContainer"] { background-color: #F7F8FC; }
[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0A0F2C 0%, #0D1B4B 60%, #0A2A6E 100%);
}
[data-testid="stSidebar"] * { color: #E8EDF7 !important; }
.dash-header {
    background: linear-gradient(135deg, #0A0F2C 0%, #0D1B4B 50%, #1a3a7c 100%);
    border-radius: 16px; padding: 28px 40px; margin-bottom: 28px;
    color: white; display: flex; align-items: center; gap: 24px;
    position: relative; overflow: hidden;
}
.dash-header h1 { font-size: 24px; font-weight: 700; margin: 0; color: white; }
.dash-header p { font-size: 13px; color: #90CDF4; margin: 4px 0 0 0; }
.badge {
    display: inline-block; background: rgba(99,179,237,0.2);
    border: 1px solid rgba(99,179,237,0.4); color: #90CDF4;
    font-size: 11px; padding: 3px 10px; border-radius: 20px; margin-top: 10px;
    font-family: 'DM Mono', monospace;
}
.kpi-card {
    background: white; border-radius: 14px; padding: 24px;
    border: 1px solid #E8EDF5; box-shadow: 0 2px 12px rgba(10,15,44,0.06);
    position: relative; overflow: hidden;
}
.kpi-card::after {
    content: ''; position: absolute; bottom: 0; left: 0; right: 0;
    height: 3px; background: var(--accent); border-radius: 0 0 14px 14px;
}
.kpi-label { font-size: 11px; font-weight: 600; text-transform: uppercase; letter-spacing: 1.2px; color: #718096; margin-bottom: 8px; }
.kpi-value { font-size: 30px; font-weight: 700; color: #0A0F2C; line-height: 1; margin-bottom: 4px; font-family: 'DM Mono', monospace; }
.kpi-sub { font-size: 12px; color: #A0AEC0; }
.section-title { font-size: 16px; font-weight: 600; color: #0A0F2C; margin: 0 0 16px 0; padding-bottom: 10px; border-bottom: 2px solid #EDF2F7; }
</style>
""", unsafe_allow_html=True)

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def img_to_base64(path):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    except:
        return None

def formatear_clp(valor):
    return f"${valor:,.0f}".replace(",", ".")

def leer_excel_informe(path):
    wb = load_workbook(path, data_only=True)
    usuarios = {}
    for nombre_hoja in wb.sheetnames:
        if nombre_hoja == "Dashboard":
            continue
        ws = wb[nombre_hoja]
        filas = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            cols = (list(row) + [None]*6)[:6]
            fecha, tipo, descripcion, monto, link, estado = cols
            if tipo not in ("Factura", "Boleta"):
                continue
            filas.append({"Fecha": fecha, "Tipo": tipo, "Descripción": descripcion,
                          "Monto": monto or 0, "Estado": estado or "—", "Usuario": nombre_hoja})
        if filas:
            usuarios[nombre_hoja] = pd.DataFrame(filas)
    return usuarios

def consolidar(usuarios):
    if not usuarios:
        return pd.DataFrame()
    return pd.concat(usuarios.values(), ignore_index=True)

# ─── SESSION STATE ────────────────────────────────────────────────────────────
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False
if "usuario" not in st.session_state:
    st.session_state.usuario = None
if "rol" not in st.session_state:
    st.session_state.rol = None
if "archivos_meses" not in st.session_state:
    st.session_state.archivos_meses = {}

LOGO_PATH = "logo_gcsp.png"

# ─── LOGIN ────────────────────────────────────────────────────────────────────
def mostrar_login():
    st.markdown("""
    <style>
    [data-testid="stAppViewContainer"] {
        background: linear-gradient(135deg, #0A0F2C 0%, #0D1B4B 55%, #0A2A6E 100%) !important;
        min-height: 100vh;
    }
    .main .block-container { padding-top: 0 !important; }

    .login-wrapper {
        display: flex; align-items: center; justify-content: center;
        min-height: 100vh; padding: 40px 20px;
    }
    .login-card {
        background: white; border-radius: 20px; padding: 48px 40px;
        width: 100%; max-width: 380px;
        box-shadow: 0 32px 80px rgba(0,0,0,0.35);
        border-top: 5px solid #3182CE;
        position: relative; overflow: hidden;
    }
    .login-card::before {
        content: ''; position: absolute; top: -60px; right: -60px;
        width: 180px; height: 180px; border-radius: 50%;
        background: radial-gradient(circle, rgba(49,130,206,0.06) 0%, transparent 70%);
    }
    .login-logo-wrap { text-align: center; margin-bottom: 20px; }
    .login-logo-wrap img {
        width: 90px; filter: drop-shadow(0 4px 12px rgba(49,130,206,0.2));
    }
    .login-title {
        text-align: center; font-size: 22px; font-weight: 700;
        color: #0A0F2C; margin-bottom: 4px; letter-spacing: -0.5px;
    }
    .login-sub {
        text-align: center; font-size: 12px; color: #A0AEC0;
        margin-bottom: 28px; text-transform: uppercase; letter-spacing: 1px;
    }
    .login-divider {
        height: 1px; background: #EDF2F7; margin-bottom: 24px;
    }
    .login-footer {
        text-align: center; margin-top: 20px;
        font-size: 11px; color: #CBD5E0;
    }
    .security-badge {
        display: inline-flex; align-items: center; gap: 6px;
        background: #F0FFF4; border: 1px solid #C6F6D5;
        color: #276749; font-size: 11px; padding: 4px 12px;
        border-radius: 20px; margin-top: 12px;
    }
    .security-dot { width: 6px; height: 6px; border-radius: 50%; background: #38A169; }
    </style>
    """, unsafe_allow_html=True)

    logo_b64 = img_to_base64(LOGO_PATH)
    logo_html = f'<img src="data:image/png;base64,{logo_b64}">' if logo_b64 else \
                '<div style="font-size:32px; font-weight:800; color:#0A0F2C;">GCS-P</div>'

    col1, col2, col3 = st.columns([1, 1.2, 1])
    with col2:
        st.markdown(f"""
        <div style='padding-top:60px;'>
        <div class="login-card">
            <div class="login-logo-wrap">{logo_html}</div>
            <div class="login-title">GCS-P Rendiciones</div>
            <div class="login-sub">Panel Ejecutivo · Acceso Seguro</div>
            <div class="login-divider"></div>
        </div>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            usuario  = st.text_input("👤 Usuario",     placeholder="Ingresa tu usuario")
            password = st.text_input("🔒 Contraseña",  type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Ingresar →", use_container_width=True)
            if submitted:
                if usuario in USUARIOS and USUARIOS[usuario]["password"] == password:
                    st.session_state.autenticado = True
                    st.session_state.usuario     = usuario
                    st.session_state.rol         = USUARIOS[usuario]["rol"]
                    st.rerun()
                else:
                    st.error("❌ Usuario o contraseña incorrectos")

        st.markdown("""
        <div style='text-align:center; margin-top:8px;'>
            <div class="security-badge">
                <div class="security-dot"></div> Conexión segura · Solo personal autorizado
            </div>
        </div>
        """, unsafe_allow_html=True)

if not st.session_state.autenticado:
    mostrar_login()
    st.stop()

# ─── SIDEBAR ──────────────────────────────────────────────────────────────────
with st.sidebar:
    logo_b64 = img_to_base64(LOGO_PATH)
    if logo_b64:
        st.markdown(f"""
        <div style='text-align:center; padding:16px 0 8px;'>
            <img src="data:image/png;base64,{logo_b64}" width="75" style="border-radius:8px; background:white; padding:4px;">
        </div>""", unsafe_allow_html=True)

    rol_badge = "👑 Admin" if st.session_state.rol == "admin" else "👁️ Gerencia"
    st.markdown(f"""
    <div style='text-align:center; padding-bottom:16px;'>
        <div style='font-size:16px; font-weight:700;'>GCS-P</div>
        <div style='font-size:11px; color:#718096;'>Panel de Rendiciones</div>
        <div style='margin-top:8px; font-size:12px; background:rgba(255,255,255,0.1);
                    border-radius:20px; padding:3px 12px; display:inline-block;'>{rol_badge}</div>
    </div>""", unsafe_allow_html=True)

    st.markdown("---")

    # Solo admin sube archivos
    if st.session_state.rol == "admin":
        st.markdown("**📤 Cargar datos por mes**")
        mes_nombre = st.text_input("Nombre del mes", placeholder="Ej: Marzo 2026")
        archivo_mes = st.file_uploader("Subir Excel del mes", type=["xlsx"])
        if st.button("➕ Agregar mes", use_container_width=True):
            if mes_nombre and archivo_mes:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    tmp.write(archivo_mes.read())
                    tmp_path = tmp.name
                usuarios_data = leer_excel_informe(tmp_path)
                os.unlink(tmp_path)
                df_mes = consolidar(usuarios_data)
                df_mes["Mes"] = mes_nombre
                st.session_state.archivos_meses[mes_nombre] = df_mes
                st.success(f"✅ {mes_nombre} agregado!")
                st.rerun()
            else:
                st.warning("Completa el nombre y sube el archivo")

        if st.session_state.archivos_meses:
            st.markdown("**📅 Meses cargados:**")
            for m in list(st.session_state.archivos_meses.keys()):
                col_m, col_x = st.columns([3, 1])
                with col_m:
                    st.markdown(f"<div style='font-size:12px; padding:3px 0;'>📅 {m}</div>", unsafe_allow_html=True)
                with col_x:
                    if st.button("✕", key=f"del_{m}"):
                        del st.session_state.archivos_meses[m]
                        st.rerun()
        st.markdown("---")

    meses_disponibles = list(st.session_state.archivos_meses.keys())
    if meses_disponibles:
        mes_sel = st.selectbox("📅 Mes", ["Todos los meses"] + meses_disponibles)
    else:
        mes_sel = "demo"

    st.markdown("---")
    if st.button("🚪 Cerrar sesión", use_container_width=True):
        for key in ["autenticado", "usuario", "rol"]:
            st.session_state[key] = None if key != "autenticado" else False
        st.rerun()

# ─── CARGAR DATOS ─────────────────────────────────────────────────────────────
DEMO = consolidar({
    "JC": pd.DataFrame([
        {"Fecha":"2026-02-03","Tipo":"Factura","Descripción":"Materiales oficina","Monto":85000,"Estado":"Aprobado","Usuario":"JC"},
        {"Fecha":"2026-02-07","Tipo":"Boleta", "Descripción":"Almuerzo reunión",  "Monto":24000,"Estado":"Aprobado","Usuario":"JC"},
        {"Fecha":"2026-02-12","Tipo":"Factura","Descripción":"Servicio técnico",  "Monto":120000,"Estado":"Pendiente","Usuario":"JC"},
    ]),
    "CV": pd.DataFrame([
        {"Fecha":"2026-02-05","Tipo":"Boleta", "Descripción":"Café proveedor",    "Monto":18000,"Estado":"Aprobado","Usuario":"CV"},
        {"Fecha":"2026-02-10","Tipo":"Factura","Descripción":"Software licencia", "Monto":210000,"Estado":"Aprobado","Usuario":"CV"},
    ]),
    "GQ": pd.DataFrame([
        {"Fecha":"2026-02-01","Tipo":"Factura","Descripción":"Impresión docs",    "Monto":45000,"Estado":"Aprobado","Usuario":"GQ"},
        {"Fecha":"2026-02-14","Tipo":"Boleta", "Descripción":"Cena trabajo",      "Monto":32000,"Estado":"Aprobado","Usuario":"GQ"},
    ]),
    "GC": pd.DataFrame([
        {"Fecha":"2026-02-08","Tipo":"Boleta", "Descripción":"Útiles escritorio", "Monto":15500,"Estado":"Aprobado","Usuario":"GC"},
        {"Fecha":"2026-02-15","Tipo":"Factura","Descripción":"Mantenimiento",     "Monto":95000,"Estado":"Aprobado","Usuario":"GC"},
    ]),
})

demo_mode = not bool(st.session_state.archivos_meses)
if demo_mode:
    df_total = DEMO.copy()
elif mes_sel == "Todos los meses":
    df_total = pd.concat(st.session_state.archivos_meses.values(), ignore_index=True)
else:
    df_total = st.session_state.archivos_meses[mes_sel].copy()

with st.sidebar:
    st.markdown("**👤 Filtrar usuario**")
    usuarios_disp = ["Todos"] + sorted(df_total["Usuario"].unique().tolist()) if not df_total.empty else ["Todos"]
    filtro_usuario = st.selectbox("", usuarios_disp, label_visibility="collapsed")
    filtro_tipo = st.selectbox("**📋 Tipo**", ["Todos", "Factura", "Boleta"])

df_filtrado = df_total.copy()
if filtro_usuario != "Todos" and not df_filtrado.empty:
    df_filtrado = df_filtrado[df_filtrado["Usuario"] == filtro_usuario]
if filtro_tipo != "Todos" and not df_filtrado.empty:
    df_filtrado = df_filtrado[df_filtrado["Tipo"] == filtro_tipo]

# ─── HEADER ───────────────────────────────────────────────────────────────────
logo_b64 = img_to_base64(LOGO_PATH)
logo_html = f'<img src="data:image/png;base64,{logo_b64}" width="60" style="border-radius:8px; background:white; padding:4px;">' if logo_b64 else ""
mes_label = mes_sel if mes_sel not in ("demo",) else "Demo"
demo_badge = '<span class="badge">⚡ DEMO</span>' if demo_mode else ''

st.markdown(f"""
<div class="dash-header">
    {logo_html}
    <div>
        <h1>Dashboard Ejecutivo de Rendiciones</h1>
        <p>Control de gastos · Facturas y Boletas · {mes_label}</p>
        <span class="badge">🔵 GCS-P</span> {demo_badge}
    </div>
</div>""", unsafe_allow_html=True)

if demo_mode and st.session_state.rol == "admin":
    st.info("💡 **Modo demo** — Carga tu primer Excel desde el panel izquierdo.", icon="📎")

# ─── KPIs ─────────────────────────────────────────────────────────────────────
total_gasto   = df_filtrado["Monto"].sum() if not df_filtrado.empty else 0
cant_facturas = int((df_filtrado["Tipo"] == "Factura").sum()) if not df_filtrado.empty else 0
cant_boletas  = int((df_filtrado["Tipo"] == "Boleta").sum()) if not df_filtrado.empty else 0
monto_fact    = df_filtrado[df_filtrado["Tipo"]=="Factura"]["Monto"].sum() if not df_filtrado.empty else 0
monto_bol     = df_filtrado[df_filtrado["Tipo"]=="Boleta"]["Monto"].sum() if not df_filtrado.empty else 0
n_usuarios    = df_filtrado["Usuario"].nunique() if not df_filtrado.empty else 0

c1,c2,c3,c4,c5 = st.columns(5)
for col, label, valor, sub, accent in [
    (c1,"GASTO TOTAL",     formatear_clp(total_gasto), f"{len(df_filtrado)} registros","#3182CE"),
    (c2,"FACTURAS",        str(cant_facturas),          formatear_clp(monto_fact),      "#E53E3E"),
    (c3,"BOLETAS",         str(cant_boletas),           formatear_clp(monto_bol),       "#38A169"),
    (c4,"MONTO FACTURAS",  formatear_clp(monto_fact),  f"{monto_fact/total_gasto*100:.1f}% del total" if total_gasto>0 else "—","#805AD5"),
    (c5,"USUARIOS ACTIVOS",str(n_usuarios),             "con rendiciones",              "#DD6B20"),
]:
    with col:
        st.markdown(f"""
        <div class="kpi-card" style="--accent:{accent}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{valor}</div>
            <div class="kpi-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)

# ─── GRÁFICOS ─────────────────────────────────────────────────────────────────
if not df_filtrado.empty:
    col_izq, col_der = st.columns([3,2])
    with col_izq:
        st.markdown('<p class="section-title">Gasto por Usuario</p>', unsafe_allow_html=True)
        resumen = df_filtrado.groupby(["Usuario","Tipo"])["Monto"].sum().reset_index()
        fig_bar = go.Figure()
        for tipo, color in [("Factura","#3182CE"),("Boleta","#38A169")]:
            d = resumen[resumen["Tipo"]==tipo]
            fig_bar.add_trace(go.Bar(name=tipo, x=d["Usuario"], y=d["Monto"],
                marker_color=color, marker_line_width=0,
                text=[formatear_clp(v) for v in d["Monto"]],
                textposition="outside", textfont=dict(size=11, color="#4A5568")))
        fig_bar.update_layout(barmode="group", paper_bgcolor="white", plot_bgcolor="white",
            font=dict(family="DM Sans", size=12), height=300,
            legend=dict(orientation="h", y=1.05, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=0,r=0,t=30,b=0),
            xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#EDF2F7", tickformat="$,.0f"))
        st.plotly_chart(fig_bar, use_container_width=True)

    with col_der:
        st.markdown('<p class="section-title">Distribución F vs B</p>', unsafe_allow_html=True)
        dist = df_filtrado.groupby("Tipo")["Monto"].sum().reset_index()
        fig_pie = go.Figure(go.Pie(
            labels=dist["Tipo"], values=dist["Monto"], hole=0.65,
            marker=dict(colors=["#3182CE","#38A169"], line=dict(color="white",width=3)),
            textinfo="label+percent", textfont=dict(size=12),
            hovertemplate="<b>%{label}</b><br>%{customdata}<extra></extra>",
            customdata=[formatear_clp(v) for v in dist["Monto"]]))
        fig_pie.add_annotation(text=f"<b>{formatear_clp(total_gasto)}</b>", x=0.5, y=0.55,
            font=dict(size=14, color="#0A0F2C", family="DM Mono"), showarrow=False)
        fig_pie.add_annotation(text="Total", x=0.5, y=0.42,
            font=dict(size=12, color="#718096"), showarrow=False)
        fig_pie.update_layout(paper_bgcolor="white", showlegend=False,
            margin=dict(l=0,r=0,t=10,b=0), height=300)
        st.plotly_chart(fig_pie, use_container_width=True)

    # Evolución mensual
    if len(st.session_state.archivos_meses) > 1 and mes_sel == "Todos los meses":
        st.markdown('<p class="section-title">📈 Evolución Mensual</p>', unsafe_allow_html=True)
        evol = [{"Mes": m,
                 "Facturas": df_m[df_m["Tipo"]=="Factura"]["Monto"].sum(),
                 "Boletas":  df_m[df_m["Tipo"]=="Boleta"]["Monto"].sum(),
                 "Total":    df_m["Monto"].sum()}
                for m, df_m in st.session_state.archivos_meses.items()]
        df_evol = pd.DataFrame(evol)
        fig_evol = go.Figure()
        fig_evol.add_trace(go.Bar(x=df_evol["Mes"], y=df_evol["Facturas"], name="Facturas", marker_color="rgba(49,130,206,0.5)"))
        fig_evol.add_trace(go.Bar(x=df_evol["Mes"], y=df_evol["Boletas"],  name="Boletas",  marker_color="rgba(56,161,105,0.5)"))
        fig_evol.add_trace(go.Scatter(x=df_evol["Mes"], y=df_evol["Total"], mode="lines+markers",
            name="Total", line=dict(color="#0A0F2C", width=2), marker=dict(size=7),
            text=[formatear_clp(v) for v in df_evol["Total"]], textposition="top center"))
        fig_evol.update_layout(barmode="stack", paper_bgcolor="white", plot_bgcolor="white",
            font=dict(family="DM Sans"), height=280,
            legend=dict(orientation="h", y=1.1, bgcolor="rgba(0,0,0,0)"),
            margin=dict(l=0,r=0,t=30,b=0),
            xaxis=dict(showgrid=False), yaxis=dict(showgrid=True, gridcolor="#EDF2F7", tickformat="$,.0f"))
        st.plotly_chart(fig_evol, use_container_width=True)

    # Resumen
    st.markdown('<p class="section-title">Resumen por Usuario</p>', unsafe_allow_html=True)
    res = df_filtrado.groupby("Usuario").agg(
        Facturas=("Tipo", lambda x: (x=="Factura").sum()),
        Boletas=("Tipo",  lambda x: (x=="Boleta").sum()),
        Total=("Monto","sum")).reset_index()
    res["Total Gastado"] = res["Total"].apply(formatear_clp)
    res["% del Total"] = (res["Total"]/res["Total"].sum()*100).apply(lambda x: f"{x:.1f}%")
    st.dataframe(res[["Usuario","Facturas","Boletas","Total Gastado","% del Total"]],
                 use_container_width=True, hide_index=True)

    # Detalle
    st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
    st.markdown('<p class="section-title">Detalle de Transacciones</p>', unsafe_allow_html=True)
    df_show = df_filtrado.copy()
    df_show["Monto ($)"] = df_show["Monto"].apply(formatear_clp)
    if "Fecha" in df_show.columns:
        df_show["Fecha"] = pd.to_datetime(df_show["Fecha"], errors="coerce").dt.strftime("%d/%m/%Y")
    cols_show = [c for c in ["Fecha","Usuario","Tipo","Descripción","Monto ($)","Estado"] if c in df_show.columns]
    st.dataframe(df_show[cols_show], use_container_width=True, hide_index=True, height=320)
else:
    st.markdown("""
    <div style='text-align:center; padding:60px; color:#A0AEC0;'>
        <div style='font-size:48px;'>📭</div>
        <div style='font-size:16px; margin-top:12px;'>Sin datos para mostrar</div>
    </div>""", unsafe_allow_html=True)

# ─── FOOTER ───────────────────────────────────────────────────────────────────
st.markdown(f"""
<div style='text-align:center; color:#A0AEC0; font-size:12px; padding:24px 16px 8px;
            border-top:1px solid #EDF2F7; font-family:DM Mono,monospace; margin-top:32px;'>
    GCS-P · Dashboard Rendiciones · {datetime.now().strftime("%d/%m/%Y %H:%M")} · {st.session_state.usuario}
</div>""", unsafe_allow_html=True)
