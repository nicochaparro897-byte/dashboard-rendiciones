import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from openpyxl import load_workbook
import os
import glob
from datetime import datetime

# ─── CONFIG ───────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Dashboard Rendiciones",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ─── ESTILOS ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] {
    font-family: 'DM Sans', sans-serif;
}

/* Fondo principal */
.main { background-color: #F7F8FC; }
[data-testid="stAppViewContainer"] { background-color: #F7F8FC; }

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(160deg, #0A0F2C 0%, #0D1B4B 60%, #0A2A6E 100%);
    border-right: none;
}
[data-testid="stSidebar"] * { color: #E8EDF7 !important; }
[data-testid="stSidebar"] .stSelectbox label { color: #A0AEC0 !important; font-size: 11px !important; text-transform: uppercase; letter-spacing: 1px; }

/* Header */
.dash-header {
    background: linear-gradient(135deg, #0A0F2C 0%, #0D1B4B 50%, #1a3a7c 100%);
    border-radius: 16px;
    padding: 32px 40px;
    margin-bottom: 28px;
    color: white;
    position: relative;
    overflow: hidden;
}
.dash-header::before {
    content: '';
    position: absolute;
    top: -50%;
    right: -10%;
    width: 400px;
    height: 400px;
    background: radial-gradient(circle, rgba(99,179,237,0.12) 0%, transparent 70%);
    border-radius: 50%;
}
.dash-header h1 {
    font-size: 28px;
    font-weight: 700;
    margin: 0;
    color: white;
    letter-spacing: -0.5px;
}
.dash-header p {
    font-size: 14px;
    color: #90CDF4;
    margin: 6px 0 0 0;
    font-weight: 400;
}
.badge {
    display: inline-block;
    background: rgba(99,179,237,0.2);
    border: 1px solid rgba(99,179,237,0.4);
    color: #90CDF4;
    font-size: 11px;
    padding: 3px 10px;
    border-radius: 20px;
    margin-top: 12px;
    font-family: 'DM Mono', monospace;
    letter-spacing: 0.5px;
}

/* KPI Cards */
.kpi-card {
    background: white;
    border-radius: 14px;
    padding: 24px;
    border: 1px solid #E8EDF5;
    box-shadow: 0 2px 12px rgba(10,15,44,0.06);
    position: relative;
    overflow: hidden;
    transition: box-shadow 0.2s ease;
}
.kpi-card:hover { box-shadow: 0 6px 24px rgba(10,15,44,0.12); }
.kpi-card::after {
    content: '';
    position: absolute;
    bottom: 0;
    left: 0;
    right: 0;
    height: 3px;
    background: var(--accent);
    border-radius: 0 0 14px 14px;
}
.kpi-label {
    font-size: 11px;
    font-weight: 600;
    text-transform: uppercase;
    letter-spacing: 1.2px;
    color: #718096;
    margin-bottom: 8px;
}
.kpi-value {
    font-size: 32px;
    font-weight: 700;
    color: #0A0F2C;
    line-height: 1;
    margin-bottom: 4px;
    font-family: 'DM Mono', monospace;
}
.kpi-sub {
    font-size: 12px;
    color: #A0AEC0;
}

/* Sección títulos */
.section-title {
    font-size: 16px;
    font-weight: 600;
    color: #0A0F2C;
    margin: 0 0 16px 0;
    padding-bottom: 10px;
    border-bottom: 2px solid #EDF2F7;
}

/* Tabla */
.stDataFrame { border-radius: 12px; overflow: hidden; }
thead tr th {
    background: #0D1B4B !important;
    color: white !important;
    font-weight: 600 !important;
    font-size: 12px !important;
    text-transform: uppercase !important;
    letter-spacing: 0.8px !important;
}
tbody tr:nth-child(even) { background: #F7F8FC; }
tbody tr:hover { background: #EBF4FF !important; }

/* Chip tipo documento */
.chip-factura {
    background: #EBF8FF;
    color: #2B6CB0;
    border: 1px solid #BEE3F8;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
}
.chip-boleta {
    background: #F0FFF4;
    color: #276749;
    border: 1px solid #C6F6D5;
    padding: 2px 10px;
    border-radius: 20px;
    font-size: 11px;
    font-weight: 600;
}

/* Separador */
.divider { height: 1px; background: #EDF2F7; margin: 24px 0; }

/* Alerta sin datos */
.no-data {
    text-align: center;
    padding: 40px;
    color: #A0AEC0;
    font-size: 14px;
}
</style>
""", unsafe_allow_html=True)


# ─── FUNCIONES ────────────────────────────────────────────────────────────────

def leer_excel_informe(path: str) -> dict:
    """Lee el archivo Informe B/F mensual y extrae datos de cada usuario."""
    wb = load_workbook(path, data_only=True)
    usuarios = {}
    hojas_excluir = {"Dashboard"}

    for nombre_hoja in wb.sheetnames:
        if nombre_hoja in hojas_excluir:
            continue
        ws = wb[nombre_hoja]
        filas = []
        for row in ws.iter_rows(min_row=8, values_only=True):
            fecha, tipo, descripcion, monto, link, estado = (list(row) + [None]*6)[:6]
            if fecha is None and tipo is None and monto is None:
                continue
            if tipo not in ("Factura", "Boleta"):
                continue
            filas.append({
                "Fecha": fecha,
                "Tipo": tipo,
                "Descripción": descripcion,
                "Monto": monto if monto else 0,
                "Link": link,
                "Estado": estado,
                "Usuario": nombre_hoja
            })
        if filas:
            usuarios[nombre_hoja] = pd.DataFrame(filas)

    return usuarios


def leer_excels_individuales(carpeta: str) -> dict:
    """Lee archivos individuales de cada usuario desde carpeta input/."""
    usuarios = {}
    archivos = glob.glob(os.path.join(carpeta, "*.xlsx"))

    for archivo in archivos:
        nombre = os.path.splitext(os.path.basename(archivo))[0]
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active

        # Buscar nombre del solicitante
        solicitante = nombre
        for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
            for i, cell in enumerate(row):
                if cell and "Solicitado por" in str(cell):
                    siguiente = row[i+1] if i+1 < len(row) else None
                    if siguiente and str(siguiente).strip() not in ("XXXX XXXX", ""):
                        solicitante = str(siguiente).strip()

        filas = []
        for row in ws.iter_rows(min_row=12, values_only=True):
            item, folio, fecha, descripcion, motivo, pagado_con, monto = (list(row) + [None]*7)[:7]
            if folio is None and monto is None:
                continue
            # Detectar tipo según contenido del folio
            tipo = "Sin Doc."
            if folio:
                texto = str(folio).lower()
                if "fact" in texto:
                    tipo = "Factura"
                elif "bolet" in texto:
                    tipo = "Boleta"
            filas.append({
                "Fecha": fecha,
                "Tipo": tipo,
                "Descripción": descripcion,
                "Monto": monto if monto else 0,
                "N° Folio": folio,
                "Usuario": solicitante
            })
        if filas:
            usuarios[solicitante] = pd.DataFrame(filas)

    return usuarios


def consolidar(usuarios: dict) -> pd.DataFrame:
    if not usuarios:
        return pd.DataFrame()
    return pd.concat(usuarios.values(), ignore_index=True)


def formatear_clp(valor: float) -> str:
    return f"${valor:,.0f}".replace(",", ".")


def color_tipo(tipo: str) -> str:
    if tipo == "Factura":
        return "🟦 Factura"
    elif tipo == "Boleta":
        return "🟩 Boleta"
    return tipo


# ─── SIDEBAR ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("""
    <div style='padding: 8px 0 24px 0;'>
        <div style='font-size:22px; font-weight:700; letter-spacing:-0.5px;'>📊 Rendiciones</div>
        <div style='font-size:11px; color:#718096; margin-top:4px; text-transform:uppercase; letter-spacing:1px;'>Panel de Control</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("**Fuente de datos**")
    modo = st.radio(
        "",
        ["📄 Informe B/F mensual", "📁 Archivos individuales"],
        label_visibility="collapsed"
    )

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    if modo == "📄 Informe B/F mensual":
        archivo = st.file_uploader("Subir Informe B/F", type=["xlsx"], key="informe")
    else:
        archivos_ind = st.file_uploader("Subir archivos de usuarios", type=["xlsx"],
                                         accept_multiple_files=True, key="individuales")

    st.markdown("<div style='height:24px'></div>", unsafe_allow_html=True)
    st.markdown("""
    <div style='font-size:11px; color:#4A5568; border-top:1px solid rgba(255,255,255,0.1); padding-top:16px;'>
        <div style='margin-bottom:6px;'>📌 Columnas esperadas:</div>
        <div style='color:#718096;'>Fecha · Tipo · Descripción</div>
        <div style='color:#718096;'>Monto · Estado</div>
    </div>
    """, unsafe_allow_html=True)


# ─── CARGAR DATOS ─────────────────────────────────────────────────────────────

usuarios = {}

if modo == "📄 Informe B/F mensual" and 'archivo' in locals() and archivo:
    import tempfile
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp.write(archivo.read())
        tmp_path = tmp.name
    usuarios = leer_excel_informe(tmp_path)
    os.unlink(tmp_path)

elif modo == "📁 Archivos individuales" and 'archivos_ind' in locals() and archivos_ind:
    import tempfile
    for arch in archivos_ind:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(arch.read())
            tmp_path = tmp.name
        wb = load_workbook(tmp_path, data_only=True)
        ws = wb.active
        nombre = os.path.splitext(arch.name)[0]
        filas = []
        for row in ws.iter_rows(min_row=12, values_only=True):
            cols = (list(row) + [None]*9)[:9]
            item, folio, fecha, desc, motivo, pagado, monto = cols[2], cols[3], cols[4], cols[5], cols[6], cols[7], cols[8]
            if folio is None and monto is None:
                continue
            tipo = "Factura" if folio and "fact" in str(folio).lower() else (
                   "Boleta" if folio and "bolet" in str(folio).lower() else "Sin Doc.")
            filas.append({"Fecha": fecha, "Tipo": tipo, "Descripción": desc,
                          "Monto": monto or 0, "Usuario": nombre})
        if filas:
            usuarios[nombre] = pd.DataFrame(filas)
        os.unlink(tmp_path)

# Usar datos de demo si no hay archivo cargado
if not usuarios:
    usuarios = {
        "JC": pd.DataFrame([
            {"Fecha": "2026-02-03", "Tipo": "Factura", "Descripción": "Materiales oficina", "Monto": 85000, "Estado": "Aprobado", "Usuario": "JC"},
            {"Fecha": "2026-02-07", "Tipo": "Boleta", "Descripción": "Almuerzo reunión", "Monto": 24000, "Estado": "Aprobado", "Usuario": "JC"},
            {"Fecha": "2026-02-12", "Tipo": "Factura", "Descripción": "Servicio técnico", "Monto": 120000, "Estado": "Pendiente", "Usuario": "JC"},
            {"Fecha": "2026-02-18", "Tipo": "Boleta", "Descripción": "Taxi cliente", "Monto": 9500, "Estado": "Aprobado", "Usuario": "JC"},
        ]),
        "CV": pd.DataFrame([
            {"Fecha": "2026-02-05", "Tipo": "Boleta", "Descripción": "Café con proveedor", "Monto": 18000, "Estado": "Aprobado", "Usuario": "CV"},
            {"Fecha": "2026-02-10", "Tipo": "Factura", "Descripción": "Software licencia", "Monto": 210000, "Estado": "Aprobado", "Usuario": "CV"},
            {"Fecha": "2026-02-20", "Tipo": "Boleta", "Descripción": "Transporte", "Monto": 7200, "Estado": "Pendiente", "Usuario": "CV"},
        ]),
        "GQ": pd.DataFrame([
            {"Fecha": "2026-02-01", "Tipo": "Factura", "Descripción": "Impresión documentos", "Monto": 45000, "Estado": "Aprobado", "Usuario": "GQ"},
            {"Fecha": "2026-02-14", "Tipo": "Boleta", "Descripción": "Cena de trabajo", "Monto": 32000, "Estado": "Aprobado", "Usuario": "GQ"},
            {"Fecha": "2026-02-22", "Tipo": "Factura", "Descripción": "Consultoría externa", "Monto": 350000, "Estado": "Pendiente", "Usuario": "GQ"},
        ]),
        "GC": pd.DataFrame([
            {"Fecha": "2026-02-08", "Tipo": "Boleta", "Descripción": "Útiles de escritorio", "Monto": 15500, "Estado": "Aprobado", "Usuario": "GC"},
            {"Fecha": "2026-02-15", "Tipo": "Factura", "Descripción": "Mantenimiento equipo", "Monto": 95000, "Estado": "Aprobado", "Usuario": "GC"},
            {"Fecha": "2026-02-25", "Tipo": "Boleta", "Descripción": "Estacionamiento", "Monto": 6000, "Estado": "Aprobado", "Usuario": "GC"},
        ]),
    }
    demo_mode = True
else:
    demo_mode = False

df_total = consolidar(usuarios)

# ─── FILTROS ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("---")
    st.markdown("**Filtros**")

    todos_usuarios = ["Todos"] + sorted(usuarios.keys())
    filtro_usuario = st.selectbox("Usuario", todos_usuarios)

    tipos_disp = ["Todos", "Factura", "Boleta"]
    filtro_tipo = st.selectbox("Tipo documento", tipos_disp)

if filtro_usuario != "Todos":
    df_filtrado = df_total[df_total["Usuario"] == filtro_usuario].copy()
else:
    df_filtrado = df_total.copy()

if filtro_tipo != "Todos":
    df_filtrado = df_filtrado[df_filtrado["Tipo"] == filtro_tipo]


# ─── HEADER ───────────────────────────────────────────────────────────────────

mes_actual = datetime.now().strftime("%B %Y").capitalize()
demo_badge = ' &nbsp;<span class="badge">⚡ DATOS DE DEMO</span>' if demo_mode else ''

st.markdown(f"""
<div class="dash-header">
    <h1>📊 Dashboard Ejecutivo de Rendiciones</h1>
    <p>Control de gastos · Facturas y Boletas · {mes_actual}</p>
    <span class="badge">🔵 EN VIVO</span>{demo_badge}
</div>
""", unsafe_allow_html=True)

if demo_mode:
    st.info("💡 **Modo demo activo** — Sube tu archivo Excel en el panel izquierdo para ver tus datos reales.", icon="📎")


# ─── KPIs ─────────────────────────────────────────────────────────────────────

total_gasto = df_filtrado["Monto"].sum()
cant_facturas = len(df_filtrado[df_filtrado["Tipo"] == "Factura"])
cant_boletas = len(df_filtrado[df_filtrado["Tipo"] == "Boleta"])
monto_facturas = df_filtrado[df_filtrado["Tipo"] == "Factura"]["Monto"].sum()
monto_boletas = df_filtrado[df_filtrado["Tipo"] == "Boleta"]["Monto"].sum()
n_usuarios = df_filtrado["Usuario"].nunique()

col1, col2, col3, col4, col5 = st.columns(5)

kpis = [
    (col1, "GASTO TOTAL", formatear_clp(total_gasto), f"{len(df_filtrado)} registros", "#3182CE"),
    (col2, "FACTURAS", f"{cant_facturas}", formatear_clp(monto_facturas), "#E53E3E"),
    (col3, "BOLETAS", f"{cant_boletas}", formatear_clp(monto_boletas), "#38A169"),
    (col4, "MONTO FACTURAS", formatear_clp(monto_facturas),
     f"{monto_facturas/total_gasto*100:.1f}% del total" if total_gasto > 0 else "—", "#805AD5"),
    (col5, "USUARIOS ACTIVOS", str(n_usuarios), "con rendiciones", "#DD6B20"),
]

for col, label, valor, sub, accent in kpis:
    with col:
        st.markdown(f"""
        <div class="kpi-card" style="--accent:{accent}">
            <div class="kpi-label">{label}</div>
            <div class="kpi-value">{valor}</div>
            <div class="kpi-sub">{sub}</div>
        </div>
        """, unsafe_allow_html=True)

st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)


# ─── GRÁFICOS ─────────────────────────────────────────────────────────────────

col_izq, col_der = st.columns([3, 2])

with col_izq:
    st.markdown('<p class="section-title">Gasto por Usuario</p>', unsafe_allow_html=True)

    resumen = df_filtrado.groupby(["Usuario", "Tipo"])["Monto"].sum().reset_index()

    colores = {"Factura": "#3182CE", "Boleta": "#38A169"}

    fig_bar = go.Figure()
    for tipo in ["Factura", "Boleta"]:
        d = resumen[resumen["Tipo"] == tipo]
        fig_bar.add_trace(go.Bar(
            name=tipo,
            x=d["Usuario"],
            y=d["Monto"],
            marker_color=colores.get(tipo, "#718096"),
            marker_line_width=0,
            text=[formatear_clp(v) for v in d["Monto"]],
            textposition="outside",
            textfont=dict(size=11, color="#4A5568"),
        ))

    fig_bar.update_layout(
        barmode="group",
        paper_bgcolor="white",
        plot_bgcolor="white",
        font=dict(family="DM Sans", size=12, color="#4A5568"),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    bgcolor="rgba(0,0,0,0)"),
        margin=dict(l=0, r=0, t=30, b=0),
        height=300,
        xaxis=dict(showgrid=False, tickfont=dict(size=12, color="#2D3748")),
        yaxis=dict(showgrid=True, gridcolor="#EDF2F7", tickformat="$,.0f",
                   tickfont=dict(size=10)),
    )
    st.plotly_chart(fig_bar, use_container_width=True)

with col_der:
    st.markdown('<p class="section-title">Distribución Facturas vs Boletas</p>', unsafe_allow_html=True)

    dist = df_filtrado.groupby("Tipo")["Monto"].sum().reset_index()

    fig_pie = go.Figure(go.Pie(
        labels=dist["Tipo"],
        values=dist["Monto"],
        hole=0.65,
        marker=dict(colors=["#3182CE", "#38A169"], line=dict(color="white", width=3)),
        textinfo="label+percent",
        textfont=dict(size=12, family="DM Sans"),
        hovertemplate="<b>%{label}</b><br>%{customdata}<extra></extra>",
        customdata=[formatear_clp(v) for v in dist["Monto"]],
    ))

    total_fmt = formatear_clp(total_gasto)
    fig_pie.add_annotation(text=f"<b>{total_fmt}</b>", x=0.5, y=0.55,
                           font=dict(size=16, color="#0A0F2C", family="DM Mono"),
                           showarrow=False)
    fig_pie.add_annotation(text="Total", x=0.5, y=0.42,
                           font=dict(size=12, color="#718096", family="DM Sans"),
                           showarrow=False)

    fig_pie.update_layout(
        paper_bgcolor="white",
        showlegend=False,
        margin=dict(l=0, r=0, t=10, b=0),
        height=300,
    )
    st.plotly_chart(fig_pie, use_container_width=True)


# ─── TABLA RESUMEN ────────────────────────────────────────────────────────────

st.markdown('<p class="section-title">Resumen por Usuario</p>', unsafe_allow_html=True)

resumen_usuarios = df_filtrado.groupby("Usuario").agg(
    Facturas=("Tipo", lambda x: (x == "Factura").sum()),
    Boletas=("Tipo", lambda x: (x == "Boleta").sum()),
    Monto_Facturas=("Monto", lambda x: x[df_filtrado.loc[x.index, "Tipo"] == "Factura"].sum()),
    Monto_Boletas=("Monto", lambda x: x[df_filtrado.loc[x.index, "Tipo"] == "Boleta"].sum()),
    Total=("Monto", "sum"),
    Registros=("Monto", "count"),
).reset_index()

resumen_usuarios["Monto Facturas"] = resumen_usuarios["Monto_Facturas"].apply(formatear_clp)
resumen_usuarios["Monto Boletas"] = resumen_usuarios["Monto_Boletas"].apply(formatear_clp)
resumen_usuarios["Total Gastado"] = resumen_usuarios["Total"].apply(formatear_clp)
resumen_usuarios["% del Total"] = (resumen_usuarios["Total"] / resumen_usuarios["Total"].sum() * 100).apply(lambda x: f"{x:.1f}%")

tabla_display = resumen_usuarios[["Usuario", "Facturas", "Boletas", "Monto Facturas", "Monto Boletas", "Total Gastado", "% del Total"]]

st.dataframe(
    tabla_display,
    use_container_width=True,
    hide_index=True,
    column_config={
        "Usuario": st.column_config.TextColumn("👤 Usuario", width="medium"),
        "Facturas": st.column_config.NumberColumn("📄 Facturas", format="%d"),
        "Boletas": st.column_config.NumberColumn("🧾 Boletas", format="%d"),
        "Monto Facturas": st.column_config.TextColumn("💰 Monto Facturas"),
        "Monto Boletas": st.column_config.TextColumn("💰 Monto Boletas"),
        "Total Gastado": st.column_config.TextColumn("🔢 Total Gastado"),
        "% del Total": st.column_config.TextColumn("📊 % del Total"),
    }
)


# ─── DETALLE ──────────────────────────────────────────────────────────────────

st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)
st.markdown('<p class="section-title">Detalle de Transacciones</p>', unsafe_allow_html=True)

df_show = df_filtrado.copy()
df_show["Monto ($)"] = df_show["Monto"].apply(formatear_clp)
if "Fecha" in df_show.columns:
    df_show["Fecha"] = pd.to_datetime(df_show["Fecha"], errors="coerce").dt.strftime("%d/%m/%Y")

cols_show = ["Fecha", "Usuario", "Tipo", "Descripción", "Monto ($)"]
if "Estado" in df_show.columns:
    cols_show.append("Estado")
cols_show = [c for c in cols_show if c in df_show.columns]

st.dataframe(
    df_show[cols_show],
    use_container_width=True,
    hide_index=True,
    height=350,
    column_config={
        "Fecha": st.column_config.TextColumn("📅 Fecha", width="small"),
        "Usuario": st.column_config.TextColumn("👤 Usuario", width="small"),
        "Tipo": st.column_config.TextColumn("📋 Tipo", width="small"),
        "Descripción": st.column_config.TextColumn("📝 Descripción"),
        "Monto ($)": st.column_config.TextColumn("💰 Monto", width="medium"),
        "Estado": st.column_config.TextColumn("✅ Estado", width="small"),
    }
)


# ─── FOOTER ───────────────────────────────────────────────────────────────────

st.markdown("<div style='height:32px'></div>", unsafe_allow_html=True)
st.markdown(f"""
<div style='text-align:center; color:#A0AEC0; font-size:12px; padding:16px;
            border-top:1px solid #EDF2F7; font-family:DM Mono, monospace;'>
    Dashboard Rendiciones · Actualizado {datetime.now().strftime("%d/%m/%Y %H:%M")}
</div>
""", unsafe_allow_html=True)
